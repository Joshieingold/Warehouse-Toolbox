from pathlib import Path

# from tkinter import *
# Explicit imports to satisfy Flake8
from tkinter import Tk, Canvas, Entry, Text, Button, PhotoImage
from pathlib import Path
from tkinter import Tk, Canvas, Entry, Text, Button, PhotoImage
import tkinter as tk
from tkinter import filedialog, messagebox, PhotoImage
import openpyxl
import clipboard
import pyautogui
import time
import subprocess
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox
import xml.etree.ElementTree as ET
import pyodbc
import subprocess
import datetime
import getpass

config_path = "c:\\Josh\\config.txt"
with open(config_path, 'r') as file:
    config = file.read().split("\n")


# Global Variables:
user = config[0]
import_speed = float(config[1])
asset_location = config[2]
database_path = config[3]
bartender_notepad = config[4]
serials_list = []
remaining_serials = 0
file_path = None

# Bom-Wip Helper
def RelativeToAssets(path: str) -> Path:
    return ASSETS_PATH / Path(path)
def OpenExcel():
    global file_path
    file_path = filedialog.askopenfilename(title="Select a File", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    if file_path:
        LoadSerials(file_path)

# Gets the serials from the Excel file.
def LoadSerials(file_path):
    global serials_list, remaining_serials
    try:
        # Opens the Excel, loads at the active sheet and clears the serials previously loaded if any.
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        serials_list.clear()
        # Gets every serial in the first column and adds it into the serial list.
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            if row[0].value is not None:
                serials_list.append(str(row[0].value))
        # Gets the amount of serials loaded and updates the display with the function.
        remaining_serials = len(serials_list)
        UpdateSerialsDisplay()
    # If serials are not able to be loaded by the file, this handles the error.
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load serials: {e}")
def UpdateSerialsDisplay():
    serials_text.delete(1.0, tk.END)
    serials_text.insert(tk.END, "\n".join(serials_list))
    count_label.config(text=f"{remaining_serials} Serials Loaded")

# Deletes all serials in the display and then puts the new ones in seperating them by new lines and updates the GUI of the serial amounts.
def UpdateSerialsDisplay():
    serials_text.delete(1.0, tk.END)
    serials_text.insert(tk.END, "\n".join(serials_list))
    count_label.config(text=f"{remaining_serials} Serials Loaded")

# Pastes the serials automatically.
def CheckPixel():
        flexLocal = pyautogui.screenshot()
        colorPixel = (flexLocal.getpixel((50,1012)))
        whitepixel = (flexLocal.getpixel((50, 1000)))
        if colorPixel != whitepixel:
            print(colorPixel)
            return False
        else: 
             return True

def PasteSerials():
    global remaining_serials
    # Handles if there are no serials in our serial list.
    if not serials_list:
        messagebox.showinfo("Info", "No serials loaded to paste.")
        return
    # Allows time for the user to focus on the target program
    time.sleep(5)  
    # For every serial in the list it is pasted and then down arrow is pressed. Serials are then removed and I attempt to make the Gui update to no avail.
    for serial in serials_list[:]:
        if serial is not None:
            isPixelGood = CheckPixel()
            while isPixelGood == False:
                time.sleep(0.75)
                isPixelGood = CheckPixel()
            if CheckPixel() == True:    
                pyautogui.typewrite(serial)
                time.sleep(0.5)
                pyautogui.hotkey("tab")
                serials_list.remove(serial)
                remaining_serials -= 1
                UpdateSerialsDisplay()
                # Time between pastes, ideally rounded down as much as possible but oracle is finicy 
                time.sleep(import_speed)
            else:
                print("Check the import for" + str(serial))
                time.sleep(2)
                pyautogui.typewrite(serial)
                time.sleep(0.5)
                pyautogui.hotkey("tab")
                serials_list.remove(serial)
                remaining_serials -= 1
                UpdateSerialsDisplay()
                # Time between pastes, ideally rounded down as much as possible but oracle is finicy 
                time.sleep(import_speed)

# Formats TV devices by reversing every 10 serials and adding the found device name to the top.
def MakeTVSheet(device):
    total_strings = len(serials_list)
    formatted_list = []

    for i in range(0, total_strings, 10):
        formatted_list.append(device)
        chunk = serials_list[i:i + 10]
        chunk.reverse()
        formatted_list.extend(chunk)

    return '\n'.join(formatted_list)

# Helper function for Modem devices, it reverses every 5 serials.
def ReverseForModems():
    total_strings = len(serials_list)
    reversed_modems = []
    for i in range(0, total_strings, 5):
        chunk = serials_list[i:i + 5][::-1]
        reversed_modems.extend(chunk)
    return reversed_modems

def MakeModemSheet(device):
    working_modems = ReverseForModems()
    total_strings = len(serials_list)
    formatted_list = []

    for i in range(0, total_strings, 8):
        formatted_list.append(device)
        chunk = working_modems[i:i + 8]
        formatted_list.extend(chunk)

    return '\n'.join(formatted_list)
# Takes the formatted string for the sheets, and executes a BAT file depending on the device that is found.
def CreatePurolatorSheet():
    global serials_list
    device = ""
    # Checks the first element in the serial list, and based on its prefix determines the device model name.
    if serials_list[0][0] == "T" and serials_list[0][1] == "M":
        device = "IPTVTCXI6HD"
    elif serials_list[0][0] == "M":
        device = "IPTVARXI6HD"
    elif serials_list[0][0] == "4" and serials_list[0][1] == "0" and serials_list[0][2] == "9":
        device = "CGM4981COM"
    elif serials_list[0][0] == "X" and serials_list[0][1] == "I" and serials_list[0][2] == "1":
        device = "SCXI11BEI"
    elif serials_list[0][0] == "3" and serials_list[0][1] == "3" and serials_list[0][2] == "6":
        device = "CGM4331COM"
    else:
        device = "TG4482A"

    # If a device is a TV device it will create the text file and execute a BAT file using it.
    if device in ["IPTVARXI6HD", "IPTVTCXI6HD", "SCXI11BEI"]:
        puro_sheet = MakeTVSheet(device)
        # Clears the stipulated notepad and then puts the purolator sheet made into it.
        with open(bartender_notepad, 'w') as file:
            file.write(f"{puro_sheet}\n")

        # CMD script changes the printer to the purolator sheet printer and then selects the modem sheet that prints 11 serials on each.
        cmd_script = """
        @echo off
        set "target_printer=55EXP_Purolator"
        powershell -Command "Get-WmiObject -Query \\"SELECT * FROM Win32_Printer WHERE ShareName='%target_printer%'\\" | Invoke-WmiMethod -Name SetDefaultPrinter"
        "C:\\Seagull\\BarTender 7.10\\Standard\\bartend.exe" /f=C:\\BTAutomation\\XI6.btw /p /x
        """
        # Updates the TEMP BAT and then executes it.
        with open("temp_cmd.bat", "w") as bat_file:
            bat_file.write(cmd_script)

        subprocess.run(["cmd.exe", "/c", "temp_cmd.bat"])
    
    # If a device is a Modem device it will create the text file and execute a BAT file using it.
    else:
        # Clears the stipulated notepad and then puts the purolator sheet made into it.
        puro_sheet = MakeModemSheet(device)
        with open(bartender_notepad, 'w') as file:
            file.write(f"{puro_sheet}\n")
        # CMD script changes the printer to the purolator sheet printer and then selects the modem sheet that prints 11 serials on each.
        cmd_script = """
        @echo off
        set "target_printer=55EXP_Purolator"
        powershell -Command "Get-WmiObject -Query \\"SELECT * FROM Win32_Printer WHERE ShareName='%target_printer%'\\" | Invoke-WmiMethod -Name SetDefaultPrinter"
        "C:\\Seagull\\BarTender 7.10\\Standard\\bartend.exe" /f=C:\\BTAutomation\\CODA.btw /p /x
        """
        # Updates the TEMP BAT and then executes it.
        with open("temp_cmd.bat", "w") as bat_file:
            bat_file.write(cmd_script)

        subprocess.run(["cmd.exe", "/c", "temp_cmd.bat"])


import pyodbc
import subprocess
import datetime
import getpass

# Define your database path and connection string

connection_string = r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=' + database_path


def CreateLaser():
    # Create the barcode file for Bartender
    with open(bartender_notepad, 'w') as file:
        for serial in serials_list:
            file.write(f"{serial}\n")

    # Prepare and write the CMD script for printing barcodes
    cmd_script = """
    @echo off
    set "target_printer=55EXP_2"
    powershell -Command "Get-WmiObject -Query \\"SELECT * FROM Win32_Printer WHERE ShareName='%target_printer%'\\" | Invoke-WmiMethod -Name SetDefaultPrinter"
    "C:\\Seagull\\BarTender 7.10\\Standard\\bartend.exe" /f=C:\\BTAutomation\\NewPrintertest.btw /p /x
    """
    
    with open("temp_cmd.bat", "w") as bat_file:
        bat_file.write(cmd_script)

    # Run the script to print barcodes
    subprocess.run(["cmd.exe", "/c", "temp_cmd.bat"])



def CreateBarcodes():
    # Create the barcode file for Bartender
    with open(bartender_notepad, 'w') as file:
        for serial in serials_list:
            file.write(f"{serial}\n")

    # Prepare and write the CMD script for printing barcodes
    cmd_script = """
    @echo off
    set "target_printer=55EXP_Barcode"
    powershell -Command "Get-WmiObject -Query \\"SELECT * FROM Win32_Printer WHERE ShareName='%target_printer%'\\" | Invoke-WmiMethod -Name SetDefaultPrinter"
    "C:\\Seagull\\BarTender 7.10\\Standard\\bartend.exe" /f=C:\\BTAutomation\\singlebar.btw /p /x
    """
    
    with open("temp_cmd.bat", "w") as bat_file:
        bat_file.write(cmd_script)

    # Run the script to print barcodes
    subprocess.run(["cmd.exe", "/c", "temp_cmd.bat"])

    # Log the data into the Access database
    try:
        # Get the current date and the username of the person running the script
        current_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        user = getpass.getuser()

        print("Connecting to the database...")
        conn = pyodbc.connect(connection_string)
        cursor = conn.cursor()

        # Insert the user, date, and each serial into the database
        for serial in serials_list:
            if len(serial) >= 2 and serial[0] == "T" and serial[1] == "M":
              device = "XI6"
            elif len(serial) >= 1 and serial[0] == "M":
                device = "XI6"
            elif len(serial) >= 3 and serial[0] == "4" and serial[1] == "0" and serial[2] == "9":
                device = "XB8"
            elif len(serial) >= 3 and serial[0] == "X" and serial[1] == "I" and serial[2] == "1":
                device = "XIONE"
            elif len(serial) >= 3 and serial[0] == "3" and serial[1] == "3" and serial[2] == "6":
                device = "XB7"
            elif len(serial) >= 2 and serial[0] == "A" and serial[1] == "S":
                device = "POD"
            else:
                device = "XB7"
            print(f"Inserting: DateLogged={current_date}, User={user}, SerialNumber={serial}, Device={device}")
            cursor.execute(
                "INSERT INTO Barcodes (DateLogged, User, SerialNumber, Device) VALUES (?, ?, ?, ?)",
                (current_date, user, serial, device)
            )

        # Commit the transaction and close the connection
        conn.commit()
        print("Data committed to the database.")
        cursor.close()
        conn.close()
        print("Connection closed.")
    
    except Exception as e:
        print(f"Error logging data to database: {e}")

# CTR Update Tool:

# Lets the GUI interact with Windows explorer and select the daily float snapshot for the use of the program.
def open_file_dialog():
    file_path = filedialog.askopenfilename(title="Select a File", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    if file_path:
        count_label.config(text=f"Selected File: {file_path}")
        process_file(file_path)

# Using Filepath (the location of the chosen excel) if the file passes being possible, The first sheet in the workbook becomes our selection.
def process_file(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        process_sheet(sheet)
    except Exception as e:
        count_label.config(text=f"Error: {str(e)}")

# Does the algorithmic work for the actual program.
def process_sheet(sheet):
    
    # Names of all the CTRS and companys in their correct order so they can be returned in correct order.
    ctr_lst = ['8009', '8017', '8037', '8038', '8041', '8047', '8052', '8067', '8080', '8093', '8975', "8986", "8990", "8994", "8997"]
    company_lst = ["NB1", "NF1"]
    double_ctr_lst = ["8982", "8993"]
    
    # Place for all data obtained in the search to be placed and counts for them to be added to for each ideration.
    data_lst = []
    combined_totals = {
        "XB8": 0, "XB7": 0, "XI6": 0, "XIONE": 0, "POD": 0, "ONT": 0, "CAMERA_1": 0, "CAMERA_2": 0, "CAMERA_3": 0
    }
    combo_totals = combined_totals.copy()

    # Iderates through every CTR name in CTR_lst
    for contractor in ctr_lst:
        
        # Gets fresh device key for each CTR, then check every piece of data in row 1 (CTR name)
        contractor_totals = {key: 0 for key in combined_totals}
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            
            # Gets the CTR name, item code, and quantity for every row
            contractor_id = str(row[1].value)
            item_code = row[3].value
            quantity = row[4].value
            
            # Checks if quantity is an integer in case some data does not factor. As well as checking if the CTR we are looking for is the one in the selected row, if so then it gets added to our totals.
            if isinstance(quantity, int) and contractor_id == contractor:
                update_totals(contractor_totals, item_code, quantity)
        
        # Once we have searched all of the snapshot for the specific CTR, that data is appended to our full data lst.
        data_lst.append(format_totals(contractor_totals))
    
    # Does the above process but combines the CTR's that are supposed to be summed and appends to lst. Importantly it's inserted into the correct location of our data lst for pasting later.
    for team in double_ctr_lst:
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            contractor_id = str(row[1].value)
            item_code = row[3].value
            quantity = row[4].value

            if isinstance(quantity, int) and contractor_id == team:
                update_totals(combo_totals, item_code, quantity)

    data_lst.insert(13, format_totals(combo_totals))

    # Does nearly the same process as aforementioned, but needs to handle data a bit differently.
    for company in company_lst:
        company_totals = {key: 0 for key in combined_totals}
        
        # Checks all rows in the sheet.
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            
            # gets the company name, and if that is the warehouse name we are looking for in the very first row, and it is in one of the effective sub-inventories we are looking for then we continue.
            company_name = str(row[0].value)
            if company_name == company and str(row[1].value) in ["S Retail N", "S Retail R", "C Retail N", "C Retail R", "C InTranst"]:
                
                # Process is the same as before.
                item_code = row[3].value
                quantity = row[4].value
                if isinstance(quantity, int):
                    update_totals(company_totals, item_code, quantity)

        data_lst.append(format_totals(company_totals))

    # Just creating a list to iderate through for based on all the data we collected that is associated with the names of the our data.
    full_list = ctr_lst + ["Combined"] + company_lst

    # Runs the automation.
    copy_data_to_excel(data_lst, full_list)

# Takes data based on the data lst we are working with for the current CTR, which item it is (so we know where to add it.), and how many items are in the float. 
def update_totals(totals, item_code, quantity):
    if item_code == "CGM4981COM":
        totals["XB8"] += quantity
    elif item_code in ["CGM4331COM", "TG4482A"]:
        totals["XB7"] += quantity
    elif item_code in ["IPTVARXI6HD", "IPTVTCXI6HD"]:
        totals["XI6"] += quantity
    elif item_code == "SCXI11BEI":
        totals["XIONE"] += quantity
    elif item_code == "XE2SGROG1":
        totals["POD"] += quantity
    elif item_code in ["XS010XB", "XS010XQ", "XS020XONT"]:
        totals["ONT"] += quantity
    elif item_code == "SCHB1AEW":
        totals["CAMERA_1"] += quantity
    elif item_code == "SCHC2AEW":
        totals["CAMERA_2"] += quantity
    elif item_code == "SCHC3AE0":
        totals["CAMERA_3"] += quantity

# Converts the data into the format we will be needing to paste in for the excel sheet (1 line per device).
def format_totals(totals):
    return '\n'.join(str(totals[key]) for key in ["XB8", "XB7", "XI6", "XIONE", "POD", "ONT", "CAMERA_1", "CAMERA_2", "CAMERA_3"])

# Iterates through our formatted data and pastes it in automated fashion.
def copy_data_to_excel(data_lst, full_list):
        for index, data in enumerate(data_lst):
            clipboard.copy(data)
            time.sleep(6)  
            pyautogui.hotkey('ctrl', 'v')
            if index < len(data_lst) - 1:
                pyautogui.hotkey("ctrl", "alt", "pagedown")
                pyautogui.hotkey('ctrl','left')


# XML Converter
def open_xml_file():
    # Open a file dialog to choose the XML file
    file_path = filedialog.askopenfilename(
        title="Select an XML File",
        filetypes=[("XML files", "*.xml")]
    )
    
    # If the user selects a file, proceed to load it
    if file_path:
        try:
            # Parse the XML file
            tree = ET.parse(file_path)
            root = tree.getroot()
            
            # Extract all <SERIAL> values
            serial_numbers = [elem.text for elem in root.findall(".//SERIAL")]
            
            if serial_numbers:
                # Insert SERIAL numbers into the text box for easy copying
                serials_text.delete(1.0, tk.END)  # Clear the text box
                serials_text.insert(tk.END, "\n".join(serial_numbers))  # Insert serials into the text box
                clipboard.copy("\n".join(serial_numbers))
            else:
                messagebox.showinfo("No SERIAL Tags Found", "No <SERIAL> tags were found in the XML file.")
                
            return serial_numbers
        except Exception as e:
            # If there is an error, show it
            messagebox.showerror("Error", f"Failed to load XML file: {e}")
            return None
    else:
        messagebox.showwarning("No file selected", "Please select a valid XML file.")
        return None

# GUI
OUTPUT_PATH = Path(__file__).parent
ASSETS_PATH = OUTPUT_PATH / Path(f"{asset_location}")


def relative_to_assets(path: str) -> Path:
    return ASSETS_PATH / Path(path)


window = Tk()

window.geometry("600x350")
window.configure(bg = "#FFFFFF")


canvas = Canvas(
    window,
    bg = "#FFFFFF",
    height = 350,
    width = 600,
    bd = 0,
    highlightthickness = 0,
    relief = "ridge"
)
# The Background box.
canvas.place(x = 0, y = 0)
canvas.create_rectangle(
    0.0,
    0.0,
    600.0,
    350.0,
    fill="#D9D9D9",
    outline="")
# The Nav bar
canvas.create_rectangle(
    0.0,
    0.0,
    600.0,
    53.0,
    fill="#FF1B1F",
    outline="")
# The Print to lazer Printer button
button_image_1 = PhotoImage(
    file=relative_to_assets("button_1.png"))
button_1 = Button(
    image=button_image_1,
    borderwidth=0,
    highlightthickness=0,
    command=lambda: CreateLaser(),
    relief="flat"
)
button_1.place(
    x=554.0,
    y=9.0,
    width=30.0,
    height=35.0
)
# The CTR Update button
button_image_2 = PhotoImage(
    file=relative_to_assets("button_2.png"))
button_2 = Button(
    image=button_image_2,
    borderwidth=0,
    highlightthickness=0,
    command=lambda: open_file_dialog(),
    relief="flat"
)
button_2.place(
    x=487.0,
    y=3.0,
    width=52.0,
    height=46.0
)
# The convert XML Button
button_image_3 = PhotoImage(
    file=relative_to_assets("button_3.png"))
button_3 = Button(
    image=button_image_3,
    borderwidth=0,
    highlightthickness=0,
    command=lambda: open_xml_file(),
    relief="flat"
)
button_3.place(
    x=433.0,
    y=3.0,
    width=54.0,
    height=46.0
)
# The print Purolator Button
button_image_4 = PhotoImage(
    file=relative_to_assets("button_4.png"))
button_4 = Button(
    image=button_image_4,
    borderwidth=0,
    highlightthickness=0,
    command=lambda: CreatePurolatorSheet(),
    relief="flat"
)
button_4.place(
    x=185.0,
    y=5.0,
    width=44.0,
    height=42.0
)
#  The print Barcodes Button
button_image_5 = PhotoImage(
    file=relative_to_assets("button_5.png"))
button_5 = Button(
    image=button_image_5,
    borderwidth=0,
    highlightthickness=0,
    command=lambda: CreateBarcodes(),
    relief="flat"
)
button_5.place(
    x=124.0,
    y=0.0,
    width=46.0,
    height=53.0
)
# The open Excel button.
button_image_6 = PhotoImage(
    file=relative_to_assets("button_6.png"))
button_6 = Button(
    image=button_image_6,
    borderwidth=0,
    highlightthickness=0,
    command=lambda: OpenExcel(),
    relief="flat"
)
button_6.place(
    x=67.0,
    y=5.0,
    width=42.48554992675781,
    height=42.0
)
# The run button
button_image_7 = PhotoImage(
    file=relative_to_assets("button_7.png"))
button_7 = Button(
    image=button_image_7,
    borderwidth=0,
    highlightthickness=0,
    command=lambda: PasteSerials(),
    relief="flat"
)
button_7.place(
    x=10.0,
    y=5.0,
    width=42.0,
    height=42.0
)

image_image_1 = PhotoImage(
    file=relative_to_assets("image_1.png"))
image_1 = canvas.create_image(
    300.0,
    71.0,
    image=image_image_1
)
# The entry Canvas
canvas.create_rectangle(
    12.0,
    90.0,
    588.0,
    335.0,
    fill="#B5B5B5",
    outline="")
# Places serials into the file.
serials_text = tk.Text(width=82, height=16, bg='#FFFFFF', fg='black',font=("Cabin", 12 * -1, "bold"))
serials_text.place(x=12,y=90)
canvas.create_text(
    255.0,
    64.0,
    anchor="nw",
    text="x Serials Loaded",
    fill="#D9D9D9",
    font=("Cabin", 12 * -1)
)

# Creates the label count for how many serials are loaded! Once again..
count_label = tk.Label(
    window,
    width=17,
    height=1,
    text="0 Serials Loaded",
    bg="#FF1B1F",
    fg="#FFFFFF",
    font=("Arial", 8, "bold")
)
count_label.place(x=224, y=60)

window.resizable(False, False)
window.mainloop()
